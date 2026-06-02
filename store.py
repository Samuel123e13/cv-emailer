"""Persistence for CV Emailer.

Stores recipients, the email draft, and SMTP settings as JSON files next to the
app. The SMTP *password* is never written to those files. If the optional
``keyring`` package is installed it is saved in the Windows Credential Manager;
otherwise it is kept in memory only for the current session.
"""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path

# Folder that holds this file -> all user data lives alongside it.
APP_DIR = Path(__file__).resolve().parent

RECIPIENTS_FILE = APP_DIR / "recipients.json"
DRAFT_FILE = APP_DIR / "draft.json"
SETTINGS_FILE = APP_DIR / "settings.json"

# Credential Manager identifiers (only used when keyring is available).
_KEYRING_SERVICE = "CV-Emailer"
# Fixed account name under which the Hunter.io API key is stored.
_HUNTER_KEY_USER = "__hunter_api_key__"

try:  # keyring is optional; the app still works without it.
    import keyring  # type: ignore

    _HAS_KEYRING = True
except Exception:  # pragma: no cover - depends on environment
    keyring = None  # type: ignore
    _HAS_KEYRING = False


# --------------------------------------------------------------------------- #
# Generic JSON helpers
# --------------------------------------------------------------------------- #
def _load_json(path: Path, default):
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def _save_json(path: Path, data) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)


# --------------------------------------------------------------------------- #
# Recipients
# --------------------------------------------------------------------------- #
# A recipient is a dict: {"name": str, "email": str, "company": str}
def load_recipients() -> list[dict]:
    data = _load_json(RECIPIENTS_FILE, [])
    return data if isinstance(data, list) else []


def save_recipients(recipients: list[dict]) -> None:
    _save_json(RECIPIENTS_FILE, recipients)


def import_recipients_from_file(path: str) -> list[dict]:
    """Read recipients from a .csv or .xlsx file.

    Columns are matched case-insensitively by header name. Recognised headers:
    name, email, company. Email is required; rows without one are skipped.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _import_xlsx(path)
    return _import_csv(path)


def _normalise_row(headers: list[str], values: list) -> dict | None:
    lookup = {h.strip().lower(): i for i, h in enumerate(headers) if h}

    def cell(*keys) -> str:
        for key in keys:
            if key in lookup:
                idx = lookup[key]
                if idx < len(values) and values[idx] is not None:
                    return str(values[idx]).strip()
        return ""

    email = cell("email", "e-mail", "email address", "mail")
    if not email:
        return None
    return {
        "name": cell("name", "full name", "contact", "first name"),
        "email": email,
        "company": cell("company", "organisation", "organization", "employer", "firm"),
    }


def _import_csv(path: str) -> list[dict]:
    out: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return out
    headers = rows[0]
    for values in rows[1:]:
        row = _normalise_row(headers, values)
        if row:
            out.append(row)
    return out


def _import_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "Reading Excel files needs the 'openpyxl' package. "
            "Install it with:  pip install openpyxl\n"
            "Or save your sheet as a .csv and import that instead."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[dict] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    headers = [str(h) if h is not None else "" for h in rows[0]]
    for values in rows[1:]:
        row = _normalise_row(headers, list(values))
        if row:
            out.append(row)
    return out


# --------------------------------------------------------------------------- #
# Draft (subject / body / attachment paths)
# --------------------------------------------------------------------------- #
def load_draft() -> dict:
    default = {"subject": "", "body": "", "attachments": []}
    data = _load_json(DRAFT_FILE, default)
    if not isinstance(data, dict):
        return default
    # Backfill any missing keys.
    for key, value in default.items():
        data.setdefault(key, value)
    return data


def save_draft(draft: dict) -> None:
    _save_json(DRAFT_FILE, draft)


# --------------------------------------------------------------------------- #
# SMTP settings (everything except the password)
# --------------------------------------------------------------------------- #
def load_settings() -> dict:
    default = {
        "host": "smtp.office365.com",
        "port": 587,
        "security": "starttls",  # "starttls" | "ssl" | "none"
        "username": "",
        "from_name": "",
        "remember_password": False,
        "send_delay_seconds": 2.0,
    }
    data = _load_json(SETTINGS_FILE, default)
    if not isinstance(data, dict):
        return default
    for key, value in default.items():
        data.setdefault(key, value)
    return data


def save_settings(settings: dict) -> None:
    # Never persist a password into the settings file.
    clean = {k: v for k, v in settings.items() if k != "password"}
    _save_json(SETTINGS_FILE, clean)


# --------------------------------------------------------------------------- #
# Password handling via Windows Credential Manager (optional)
# --------------------------------------------------------------------------- #
def has_secure_storage() -> bool:
    return _HAS_KEYRING


def save_password(username: str, password: str) -> bool:
    if not (_HAS_KEYRING and username):
        return False
    try:
        keyring.set_password(_KEYRING_SERVICE, username, password)
        return True
    except Exception:
        return False


def load_password(username: str) -> str:
    if not (_HAS_KEYRING and username):
        return ""
    try:
        return keyring.get_password(_KEYRING_SERVICE, username) or ""
    except Exception:
        return ""


def delete_password(username: str) -> None:
    if not (_HAS_KEYRING and username):
        return
    try:
        keyring.delete_password(_KEYRING_SERVICE, username)
    except Exception:
        pass


# --------------------------------------------------------------------------- #
# Hunter.io API key (stored in Credential Manager, never in a file)
# --------------------------------------------------------------------------- #
def save_hunter_key(key: str) -> bool:
    if not _HAS_KEYRING:
        return False
    try:
        if key:
            keyring.set_password(_KEYRING_SERVICE, _HUNTER_KEY_USER, key)
        else:
            keyring.delete_password(_KEYRING_SERVICE, _HUNTER_KEY_USER)
        return True
    except Exception:
        return False


def load_hunter_key() -> str:
    if not _HAS_KEYRING:
        return ""
    try:
        return keyring.get_password(_KEYRING_SERVICE, _HUNTER_KEY_USER) or ""
    except Exception:
        return ""
